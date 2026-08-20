#!/usr/bin/env python3
"""
Extract V3 venue data from Blaniko_Venues_FINAL_v3.xlsx → JSON payload.

Reads sheet "Blaniko Official MVP" and produces a validated JSON file
suitable for the replace_venues_v3 RPC or the TS dry-run importer.

Usage:
    python3 scripts/extract_v3_to_json.py [--xlsx PATH] [--output PATH]

Workbook path resolution (highest priority first):
    1. --xlsx PATH                 explicit CLI argument
    2. BLANIKO_V3_XLSX env var     explicit override
    3. ~/Desktop/Blaniko/Excel file/Blaniko_Venues_FINAL_v3.xlsx  (portable default)

    The resolved path is printed. If it does not exist the script fails loudly
    and NEVER silently falls back to another (older V2/V3/backup) workbook.

Defaults:
    --output scripts/v3_venues_payload.json
"""

import argparse
import json
import os
import re
import sys
import unicodedata
from datetime import date, datetime

import openpyxl

# ─── Constants ────────────────────────────────────────────────────────────────

SHEET_NAME = "Blaniko Official MVP"

# Permanently retired venues — removed from the MVP and NEVER to be reintroduced.
# Rows with these IDs are skipped during extraction so a regeneration from the
# source xlsx can never bring them back. Their BLK IDs must never be reassigned.
#   BLK-0020  E-Blue Gaming Center                    (permanently closed)
#   BLK-0045  FCC football sidi maarouf كرة القدم…    (product/editorial decision)
RETIRED_IDS = {"BLK-0020", "BLK-0045"}

EXPECTED_COUNT = 97

EXPECTED_IDS = (
    set(
        [f"BLK-{i:04d}" for i in range(1, 37)]
        + [f"BLK-{i:04d}" for i in range(38, 101)]
    )
    - RETIRED_IDS
)

# BLK-0045 (FCC) formerly had an Unknown contact but is now retired/excluded;
# BLK-0044 (EFA, now active) has a known phone, so the active set has 3 Unknowns.
UNKNOWN_CONTACT_IDS = {"BLK-0068", "BLK-0076", "BLK-0089"}

VALID_INDOOR_OUTDOOR = {"Indoor", "Outdoor", "Indoor / Outdoor"}

# Column indices (1-based, matching the xlsx)
COL_ID = 1
COL_NAME = 2
COL_GOOGLE_MAPS = 3
COL_CATEGORY = 4
COL_SUBCATEGORY = 5
COL_ADDITIONAL_EXP = 6
COL_LOCATION = 7
COL_CONTACT = 8
COL_AUDIENCE = 9
COL_EXPERIENCE_DESC = 10
COL_ATMOSPHERE = 11
COL_INDOOR_OUTDOOR = 12
COL_WEBSITE = 13
COL_FACEBOOK = 14
COL_INSTAGRAM = 15
COL_OPENING_HOURS = 16
COL_BOOKING_METHOD = 17
COL_BOOKING_LINK = 18
COL_LAST_VERIFIED = 19
COL_STATUS = 20
COL_PRICE = 21
COL_PRICE_DETAILS = 22
COL_VERIFIED_BY = 23
# Manually curated main area / quartier for discovery filtering (Phase 1 quartier
# migration). Feeds the payload `neighborhood` field, which the backend composes
# into the frontend `area`. Distinct from COL_LOCATION (full street address).
COL_QUARTIER = 25


# ─── Helpers ──────────────────────────────────────────────────────────────────


def slugify(text: str) -> str:
    """Slugify a venue name: normalize accents, lowercase, replace non-alnum with hyphens."""
    # Normalize unicode accents to ASCII
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_text = nfkd.encode("ascii", "ignore").decode("ascii")
    # Lowercase
    slug = ascii_text.lower()
    # Replace non-alphanumeric with hyphens
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    # Strip leading/trailing hyphens
    slug = slug.strip("-")
    # Collapse double hyphens (safety)
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug


def make_slug(name: str, external_id: str) -> str:
    """Generate slug: slugified-name-blk-NNNN"""
    slug = slugify(name) + "-" + external_id.lower()
    # Final safety: no double hyphens
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug


def split_comma_list(value: str) -> list[str]:
    """Split a comma-separated string into trimmed, non-empty items."""
    items = [item.strip() for item in value.split(",")]
    return [item for item in items if item]


def split_additional_experiences(value: str) -> list[str] | None:
    """
    Split Additional Experiences using the exact rule:
    - "Unknown" → None
    - Contains newline → split by newline
    - Otherwise → split by comma
    Then trim, remove empty, deduplicate (preserving order).
    """
    if value == "Unknown":
        return None

    if "\n" in value:
        items = value.split("\n")
    else:
        items = value.split(",")

    # Trim and remove empty
    items = [item.strip() for item in items]
    items = [item for item in items if item]

    # Deduplicate preserving order
    seen: set[str] = set()
    deduped: list[str] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            deduped.append(item)

    return deduped if deduped else None


def unknown_to_none(value: str | None) -> str | None:
    """Convert 'Unknown' to None for fields where that mapping applies."""
    if value is None or str(value).strip() == "Unknown":
        return None
    v = str(value).strip()
    return v if v else None


def get_hyperlink_target(cell) -> str | None:
    """Get the hyperlink target from a cell, falling back to cell value."""
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    return None


def is_google_maps_url(url: str) -> bool:
    """Validate that a URL is a Google Maps link (HTTPS, Google-owned domain)."""
    if not url.startswith("https://"):
        return False
    # Accept various Google Maps URL forms
    google_maps_patterns = [
        "google.com/maps",
        "google.co",
        "maps.google",
        "goo.gl/maps",
        "maps.app.goo.gl",
    ]
    url_lower = url.lower()
    return any(pattern in url_lower for pattern in google_maps_patterns)


def safe_date(value) -> str | None:
    """Convert a datetime or date to ISO date string safely."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    # Try parsing string
    s = str(value).strip()
    if not s or s == "Unknown":
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return None


def split_booking_method(value: str) -> list[str] | None:
    """Split booking method. Unknown → None."""
    if value == "Unknown":
        return None
    items = split_comma_list(value)
    return items if items else None


# ─── Main extraction ─────────────────────────────────────────────────────────


def extract(xlsx_path: str) -> list[dict]:
    """Extract all 97 venues from the V3 xlsx file."""

    if not os.path.exists(xlsx_path):
        print(f"ERROR: File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path)

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[SHEET_NAME]

    venues: list[dict] = []
    errors: list[str] = []

    for row_num in range(2, ws.max_row + 1):
        external_id = ws.cell(row=row_num, column=COL_ID).value
        if external_id is None:
            break

        external_id = str(external_id).strip()
        if not external_id.startswith("BLK-"):
            break

        # Skip permanently retired venues so they are never reintroduced.
        if external_id in RETIRED_IDS:
            continue

        name = ws.cell(row=row_num, column=COL_NAME).value
        if name is None:
            errors.append(f"Row {row_num}: missing name for {external_id}")
            continue
        name = str(name).strip()

        # Google Maps — use hyperlink target
        maps_cell = ws.cell(row=row_num, column=COL_GOOGLE_MAPS)
        google_maps_url = get_hyperlink_target(maps_cell)
        if google_maps_url is None:
            errors.append(f"{external_id}: no Google Maps hyperlink target")
        elif not is_google_maps_url(google_maps_url):
            errors.append(f"{external_id}: invalid Google Maps URL: {google_maps_url}")

        # Category / Subcategory
        category = str(ws.cell(row=row_num, column=COL_CATEGORY).value or "").strip()
        subcategory = ws.cell(row=row_num, column=COL_SUBCATEGORY).value
        subcategory = str(subcategory).strip() if subcategory else None

        # Additional Experiences
        ae_raw = ws.cell(row=row_num, column=COL_ADDITIONAL_EXP).value
        additional_experiences = None
        if ae_raw is not None:
            ae_str = str(ae_raw).strip()
            if ae_str:
                additional_experiences = split_additional_experiences(ae_str)

        # Location (full street address)
        location_text = ws.cell(row=row_num, column=COL_LOCATION).value
        location_text = str(location_text).strip() if location_text else None

        # Main area / quartier (curated) → payload neighborhood → frontend area.
        quartier = ws.cell(row=row_num, column=COL_QUARTIER).value
        quartier = str(quartier).strip() if quartier else None

        # Contact information — MUST be a string, preserve exactly
        contact_cell = ws.cell(row=row_num, column=COL_CONTACT)
        contact_raw = contact_cell.value
        if not isinstance(contact_raw, str):
            errors.append(
                f"{external_id}: Contact information is not a string "
                f"(type={type(contact_raw).__name__}, value={contact_raw!r}). "
                f"FAILING extraction."
            )
            print(f"FATAL: {external_id} contact is not a string: {contact_raw!r}", file=sys.stderr)
            sys.exit(1)
        contact_information = contact_raw  # Preserve exactly, no normalization

        # Audience → array
        audience_raw = ws.cell(row=row_num, column=COL_AUDIENCE).value
        audience_tags = split_comma_list(str(audience_raw)) if audience_raw else []

        # Experience description
        exp_desc = ws.cell(row=row_num, column=COL_EXPERIENCE_DESC).value
        experience_description = str(exp_desc).strip() if exp_desc else None

        # Atmosphere → array
        atmo_raw = ws.cell(row=row_num, column=COL_ATMOSPHERE).value
        atmosphere_tags = split_comma_list(str(atmo_raw)) if atmo_raw else []

        # Indoor/Outdoor
        io_raw = ws.cell(row=row_num, column=COL_INDOOR_OUTDOOR).value
        indoor_outdoor = str(io_raw).strip() if io_raw else None
        if indoor_outdoor and indoor_outdoor not in VALID_INDOOR_OUTDOOR:
            errors.append(f"{external_id}: invalid Indoor/Outdoor value: {indoor_outdoor!r}")

        # Website — Unknown → None
        website = unknown_to_none(ws.cell(row=row_num, column=COL_WEBSITE).value)

        # Facebook — use hyperlink target if available, Unknown → None
        fb_cell = ws.cell(row=row_num, column=COL_FACEBOOK)
        facebook = None
        if fb_cell.hyperlink and fb_cell.hyperlink.target:
            facebook = fb_cell.hyperlink.target
        elif fb_cell.value and str(fb_cell.value).strip() != "Unknown":
            facebook = str(fb_cell.value).strip() or None

        # Instagram — use hyperlink target if available, Unknown → None
        ig_cell = ws.cell(row=row_num, column=COL_INSTAGRAM)
        instagram = None
        if ig_cell.hyperlink and ig_cell.hyperlink.target:
            instagram = ig_cell.hyperlink.target
        elif ig_cell.value and str(ig_cell.value).strip() != "Unknown":
            instagram = str(ig_cell.value).strip() or None

        # Opening Hours — preserve exactly
        hours_raw = ws.cell(row=row_num, column=COL_OPENING_HOURS).value
        opening_hours_raw = str(hours_raw).strip() if hours_raw else None

        # Booking Method — Unknown → None
        bm_raw = ws.cell(row=row_num, column=COL_BOOKING_METHOD).value
        booking_method = split_booking_method(str(bm_raw).strip()) if bm_raw else None

        # Booking Link — Unknown → None
        bl_raw = ws.cell(row=row_num, column=COL_BOOKING_LINK).value
        booking_link = unknown_to_none(bl_raw)

        # Last Verified Date
        last_verified_date = safe_date(ws.cell(row=row_num, column=COL_LAST_VERIFIED).value)

        # Verified By
        verified_by = ws.cell(row=row_num, column=COL_VERIFIED_BY).value
        verified_by = str(verified_by).strip() if verified_by else None

        # Generate slug
        slug = make_slug(name, external_id)

        # Build venue object
        venue: dict = {
            "external_id": external_id,
            "name": name,
            "slug": slug,
            "category": category,
            "subcategory": subcategory,
            "additional_experiences": additional_experiences,
            "location_text": location_text,
            "google_maps_url": google_maps_url,
            "contact_information": contact_information,
            "audience_tags": audience_tags,
            "experience_description": experience_description,
            "atmosphere_tags": atmosphere_tags,
            "indoor_outdoor": indoor_outdoor,
            "website": website,
            "facebook": facebook,
            "instagram": instagram,
            "opening_hours_raw": opening_hours_raw,
            "booking_method": booking_method,
            "booking_link": booking_link,
            "last_verified_date": last_verified_date,
            "verified_by": verified_by,
            # Trust / verification — explicitly assigned during V3 import
            "research_status": "Verified",
            "verification_level": "publicly_researched",
            # Price fields — always NULL
            "price": None,
            "price_details": None,
            "price_level": None,
            # Coordinates — always NULL
            "lat": None,
            "lng": None,
            # Active
            "is_active": True,
            "source": "v3_import",
            # ─── Compatibility fields ──────────────────────────────────────
            "address": location_text,
            "phone": contact_information,  # Exactly equal to contact_information
            "google_maps_link": google_maps_url,  # Exactly equal to google_maps_url
            "audience": ", ".join(audience_tags) if audience_tags else None,
            "vibe": ", ".join(atmosphere_tags) if atmosphere_tags else None,
            "short_description": experience_description,
            "overview": experience_description,
            "region": None,
            "neighborhood": quartier,
            "image_url": None,
        }

        venues.append(venue)

    if errors:
        print(f"\n{'='*60}", file=sys.stderr)
        print(f"EXTRACTION ERRORS ({len(errors)}):", file=sys.stderr)
        for e in errors:
            print(f"  - {e}", file=sys.stderr)
        print(f"{'='*60}", file=sys.stderr)

    return venues


# ─── Validation ───────────────────────────────────────────────────────────────


def validate(venues: list[dict]) -> bool:
    """Run all payload validations. Returns True if valid."""
    errors: list[str] = []

    # Exactly 97 rows
    if len(venues) != EXPECTED_COUNT:
        errors.append(f"Expected {EXPECTED_COUNT} venues, got {len(venues)}")

    # Exact ID set
    actual_ids = {v["external_id"] for v in venues}
    if actual_ids != EXPECTED_IDS:
        missing = EXPECTED_IDS - actual_ids
        extra = actual_ids - EXPECTED_IDS
        if missing:
            errors.append(f"Missing IDs: {sorted(missing)}")
        if extra:
            errors.append(f"Extra IDs: {sorted(extra)}")

    # BLK-0037 absent
    if "BLK-0037" in actual_ids:
        errors.append("BLK-0037 must be absent")

    # 97 unique external IDs
    ext_ids = [v["external_id"] for v in venues]
    if len(ext_ids) != len(set(ext_ids)):
        errors.append("Duplicate external IDs found")

    # 97 unique slugs
    slugs = [v["slug"] for v in venues]
    if len(slugs) != len(set(slugs)):
        errors.append("Duplicate slugs found")

    # BLK-0038 identity
    blk38 = next((v for v in venues if v["external_id"] == "BLK-0038"), None)
    if blk38:
        if blk38["name"] != "OASIS SPORTS CITY":
            errors.append(f"BLK-0038 expected name 'OASIS SPORTS CITY', got '{blk38['name']}'")
    else:
        errors.append("BLK-0038 not found")

    # BLK-0100 identity
    blk100 = next((v for v in venues if v["external_id"] == "BLK-0100"), None)
    if blk100:
        expected_name = "LE M SPA \u2013 Hammam & Massage Spa Casablanca"
        if blk100["name"] != expected_name:
            errors.append(f"BLK-0100 expected name '{expected_name}', got '{blk100['name']}'")
    else:
        errors.append("BLK-0100 not found")

    # 97 Google Maps hyperlink targets
    maps_count = sum(1 for v in venues if v["google_maps_url"])
    if maps_count != 97:
        errors.append(f"Expected 97 Google Maps URLs, got {maps_count}")

    # 97 Contact information strings
    contacts = [v["contact_information"] for v in venues]
    string_contacts = [c for c in contacts if isinstance(c, str)]
    if len(string_contacts) != 97:
        errors.append(f"Expected 97 string contacts, got {len(string_contacts)}")

    # 3 Unknown contacts
    unknown_contacts = [v for v in venues if v["contact_information"] == "Unknown"]
    if len(unknown_contacts) != 3:
        errors.append(f"Expected 3 Unknown contacts, got {len(unknown_contacts)}")

    # Correct Unknown contact IDs
    unknown_ids = {v["external_id"] for v in unknown_contacts}
    if unknown_ids != UNKNOWN_CONTACT_IDS:
        errors.append(f"Wrong Unknown contact IDs: expected {UNKNOWN_CONTACT_IDS}, got {unknown_ids}")

    # 94 non-Unknown contacts
    non_unknown = [c for c in contacts if c != "Unknown"]
    if len(non_unknown) != 94:
        errors.append(f"Expected 94 non-Unknown contacts, got {len(non_unknown)}")

    # 0 null contacts
    null_contacts = [c for c in contacts if c is None]
    if len(null_contacts) != 0:
        errors.append(f"Expected 0 null contacts, got {len(null_contacts)}")

    # Compatibility: phone == contact_information
    for v in venues:
        if v["phone"] != v["contact_information"]:
            errors.append(f"{v['external_id']}: phone != contact_information")

    # Audience arrays non-empty
    for v in venues:
        if not v["audience_tags"] or len(v["audience_tags"]) == 0:
            errors.append(f"{v['external_id']}: empty audience_tags")

    # Atmosphere arrays non-empty
    for v in venues:
        if not v["atmosphere_tags"] or len(v["atmosphere_tags"]) == 0:
            errors.append(f"{v['external_id']}: empty atmosphere_tags")

    # Experience descriptions non-empty
    for v in venues:
        if not v["experience_description"]:
            errors.append(f"{v['external_id']}: empty experience_description")

    # Indoor/Outdoor only approved values
    for v in venues:
        if v["indoor_outdoor"] and v["indoor_outdoor"] not in VALID_INDOOR_OUTDOOR:
            errors.append(f"{v['external_id']}: invalid indoor_outdoor: {v['indoor_outdoor']}")

    # Price fields all NULL
    for v in venues:
        if v["price"] is not None:
            errors.append(f"{v['external_id']}: price must be null")
        if v["price_details"] is not None:
            errors.append(f"{v['external_id']}: price_details must be null")
        if v["price_level"] is not None:
            errors.append(f"{v['external_id']}: price_level must be null")

    # lat/lng NULL
    for v in venues:
        if v["lat"] is not None:
            errors.append(f"{v['external_id']}: lat must be null")
        if v["lng"] is not None:
            errors.append(f"{v['external_id']}: lng must be null")

    # research_status / verification_level
    for v in venues:
        if v["research_status"] != "Verified":
            errors.append(f"{v['external_id']}: research_status must be 'Verified'")
        if v["verification_level"] != "publicly_researched":
            errors.append(f"{v['external_id']}: verification_level must be 'publicly_researched'")

    # is_active = true
    for v in venues:
        if v["is_active"] is not True:
            errors.append(f"{v['external_id']}: is_active must be true")

    # No slug contains "--"
    for v in venues:
        if "--" in v["slug"]:
            errors.append(f"{v['external_id']}: slug contains '--': {v['slug']}")

    # Print results
    if errors:
        print(f"\n{'='*60}")
        print(f"VALIDATION FAILED — {len(errors)} error(s):")
        for e in errors:
            print(f"  ✗ {e}")
        print(f"{'='*60}")
        return False
    else:
        print(f"\n{'='*60}")
        print("VALIDATION PASSED — all checks OK")
        print(f"  ✓ {len(venues)} venues extracted")
        print(f"  ✓ Exact ID set (BLK-0001..BLK-0036, BLK-0038..BLK-0100, minus retired BLK-0020/BLK-0045)")
        print(f"  ✓ BLK-0037 absent; BLK-0020 & BLK-0045 retired")
        print(f"  ✓ 97 unique external IDs")
        print(f"  ✓ 97 unique slugs")
        print(f"  ✓ BLK-0038 = OASIS SPORTS CITY")
        print(f"  ✓ BLK-0100 = LE M SPA identity confirmed")
        print(f"  ✓ 97 Google Maps hyperlink targets")
        print(f"  ✓ 97 contact strings (3 Unknown, 94 non-Unknown, 0 null)")
        print(f"  ✓ Unknown contacts: {sorted(UNKNOWN_CONTACT_IDS)}")
        print(f"  ✓ All audience arrays non-empty")
        print(f"  ✓ All atmosphere arrays non-empty")
        print(f"  ✓ All experience descriptions non-empty")
        print(f"  ✓ Indoor/Outdoor values valid")
        print(f"  ✓ price/price_details/price_level all NULL")
        print(f"  ✓ lat/lng all NULL")
        print(f"  ✓ research_status=Verified, verification_level=publicly_researched")
        print(f"  ✓ is_active=true on all venues")
        print(f"  ✓ No slugs contain '--'")
        print(f"{'='*60}")
        return True


# ─── CLI ──────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="Extract V3 venues from xlsx to JSON")
    parser.add_argument(
        "--xlsx",
        default=None,
        help="Path to the V3 xlsx file (overrides BLANIKO_V3_XLSX and the default)",
    )
    parser.add_argument(
        "--output",
        default=os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "v3_venues_payload.json",
        ),
        help="Output JSON file path",
    )
    args = parser.parse_args()

    # Resolve the workbook path portably (never hardcode a username):
    #   1. --xlsx  2. BLANIKO_V3_XLSX env  3. ~/Desktop/Blaniko/Excel file/…
    default_xlsx = os.path.join(
        os.path.expanduser("~"),
        "Desktop",
        "Blaniko",
        "Excel file",
        "Blaniko_Venues_FINAL_v3.xlsx",
    )
    if args.xlsx:
        xlsx_path = args.xlsx
        xlsx_source = "--xlsx argument"
    elif os.environ.get("BLANIKO_V3_XLSX"):
        xlsx_path = os.environ["BLANIKO_V3_XLSX"]
        xlsx_source = "BLANIKO_V3_XLSX env var"
    else:
        xlsx_path = default_xlsx
        xlsx_source = "portable default (~/Desktop/Blaniko/Excel file/…)"

    print(f"Workbook path ({xlsx_source}): {xlsx_path}")

    # Fail loudly rather than silently reading the wrong (older) workbook.
    if not os.path.isfile(xlsx_path):
        print(
            f"\nERROR: V3 workbook not found at resolved path:\n  {xlsx_path}\n"
            "Provide the correct workbook via --xlsx PATH or the BLANIKO_V3_XLSX "
            "environment variable. This script never falls back to another spreadsheet.",
            file=sys.stderr,
        )
        sys.exit(1)

    print(f"Reading: {xlsx_path}")
    venues = extract(xlsx_path)

    print(f"\nExtracted {len(venues)} venues")

    valid = validate(venues)

    # Write output regardless (for inspection), but warn if invalid
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(venues, f, indent=2, ensure_ascii=False)
    print(f"\nWritten to: {args.output}")

    if not valid:
        print("\n⚠  Payload has validation errors — review before import!", file=sys.stderr)
        sys.exit(1)

    print("\n✓ Payload ready for dry-run import.")


if __name__ == "__main__":
    main()
